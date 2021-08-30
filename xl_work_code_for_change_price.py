import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def worksheet_process(filename):
    wb = xl.load_workbook(filename) #optional, you directly put the name of the file
    sheet_no = input("Enter the sheet number: ")
    sheet = wb[f'Sheet{sheet_no}']
    #cell = sheet.cell(1,1)
    # print(cell.value)
    # print(sheet.max_row) #maximum rows

    col_name = input("Enter the new column name: ") # optional
    for row in range(2, (sheet.max_row) +1):
        cell = sheet.cell(row,3)
        # print(cell.value)
        
        # for naming the colomn-4        # optional
        col4_row1 = sheet.cell(1,4)
        col4_row1.value = col_name

        #for changing values
        corrected_price =  cell.value*0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price


    #adding chart for the new values
    choice = input("Do you want to draw the graph of it (Yes/No)?").lower()
    if choice == 'yes':
        values = Reference(sheet,
                    min_row=2,
                    max_row=sheet.max_row,
                    min_col=4,
                    max_col=4)
        chart = BarChart()
        chart.add_data(values)
        sheet.add_chart(chart,'e2')
        print("Graph has drawn!")
    elif choice == 'no':
        print("Thanks")
    #save file
    wb.save('newfile.xlsx') #if you just to save in existing file then leave that as filename
    print("You are done!")

filename = input("Enter your file name in the current directory: ")
worksheet_process(filename)